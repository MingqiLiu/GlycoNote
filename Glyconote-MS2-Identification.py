print ('\nGlycoNote is made by Mingqi Liu, Guy Treves, Andres Guerrero working under professor Carlito Lebrilla. University of California, Davis')

###
### import module

import os
import time
import random
import warnings
import itertools
import matplotlib.pyplot as Spe_plt
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import Workbook

###
### variable

# testing parameters
Decoy_repeat_max_time = 7200
#Decoy_repeat_max_time = 1000

Decoy_vs_target_repeat = 200

Test_random_range = 1.0

FDR_threshold_int = 0.02
FDR_threshold_seq = 0.02
#FDR_threshold_int = 2
#FDR_threshold_seq = 2

Decoy_repeat_max_round = 10000
Para_decoy_round = 10000

Test_charge_annotation = 3
Test_charge_max = 3

Test_min_spe = 50
Test_max_group = 10

Para_filter_sn = 0
Para_filter_top = 50

Test_output_spe = False
Test_output_spe = True

Test_output_distribution = False
#Test_output_distribution = True

Test_output_FDR = False
#Test_output_FDR = True

Test_output_decoy_list = False
#Test_output_decoy_list = True

Test_output_dia_MGF = False
#Test_output_dia_MGF = True

Para_filename = filedialog.askopenfilename (initialdir = (os.getcwd()), filetypes = [('MS/MS data', '.mgf')], title = ('Select MS/MS data'))
Parameter_file = ''

'''
if 'HMO' in Para_filename:
	Parameter_file = 'C:/Glyconote-MS2-Identification-Input-HMO.txt'
if 'NGlycan' in Para_filename:
	Parameter_file = 'C:/Glyconote-MS2-Identification-Input-NGlycan.txt'
if 'OGlycan' in Para_filename:
	Parameter_file = 'C:/Glyconote-MS2-Identification-Input-OGlycan.txt'
if 'Poly' in Para_filename:
	Parameter_file = 'C:/Glyconote-MS2-Identification-Input-Poly.txt'
if 'txt' not in Parameter_file:
	Parameter_file = filedialog.askopenfilename (initialdir = (os.getcwd()), filetypes = [('parameter file', '.txt')], title = ('Select parameter file'))
'''

Parameter_file = filedialog.askopenfilename (initialdir = (os.getcwd()), filetypes = [('parameter file', '.txt')], title = ('Select parameter file'))
Parameter_file = os.path.abspath(Parameter_file)

### variable of parameters
Time_start = time.time()
warnings.filterwarnings('ignore')

Para_name_array = []
Para_var_array = []

Para_tolerance_precursor = []
Para_tolerance_precursor_unit = []
Para_tolerance_fragment = []
Para_tolerance_fragment_unit = []

Para_charged_mass = []
Para_temfalse = []
Para_c13 = []
Para_NGlycan = []

Para_filter_diagnostic_ions = []
Para_filter_diagnostic_ions_top = 5
Para_filter_diagnostic_ions_count = 2

Para_composition_glycan = ''

### variable of MS fragmentation

MS2_spectra_no = 0
MS2_precursor_mass_max = 5000

MS2_title = []
MS2_charge = []
MS2_pepmass = []
MS2_rt = []

MS2_ingredient_array = []
MS2_fragment_array = []

MS2_real_mass_exp = []
MS2_real_abundance = []
MS2_real_mass_theo = []
MS2_real_desc = []
MS2_real_comp = []

MS2_false_mass_exp = []
MS2_false_abundance = []
MS2_false_mass_theo = []
MS2_false_desc = []
MS2_false_comp = []

MS2_nonassignable = []

MS2_intensity_sum_real = 0
MS2_intensity_sum_false = 0

MS2_diag_desc = []
MS2_diag_mass = []
MS2_diag_mass_orig = []
MS2_diag_exist = []
MS2_diag_exist_orig = []

MGF_list_diagnostic_ions = []
MGF_list_annotated = []
MGF_list_nonassignable = []

File_result = []
File_result_decoy = []
File_result_target_int = []
File_result_target_seq = []
File_result_fragment = []

### variable of spectral annotation
Spe_x_list = []
Spe_y_list = []
Spe_y_list_counter = 0
Spe_annotate_list = []
Spe_plot_allpeak_mass = []
Spe_plot_allpeak_abundance = []
Spe_counter = 1

###
### class

class ingredient:
	def __init__ (self, name, mass, occur_min, occur_max):
		self.name = name
		self.mass = mass
		self.occur_min = occur_min
		self.occur_max = occur_max

class combination:
	def __init__ (self, description, composition, mass, charge, count):
		self.description = description
		self.composition = composition
		self.mass = mass
		self.charge = charge
		self.count = count

###
### function

### Thread priority of program
def Func_setpriority (pid = None, priority = 0):

	import win32api,win32process,win32con

	priorityclasses = [win32process.IDLE_PRIORITY_CLASS,
                       win32process.BELOW_NORMAL_PRIORITY_CLASS,
                       win32process.NORMAL_PRIORITY_CLASS,
                       win32process.ABOVE_NORMAL_PRIORITY_CLASS,
                       win32process.HIGH_PRIORITY_CLASS,
                       win32process.REALTIME_PRIORITY_CLASS]
	if pid == None:
		pid = win32api.GetCurrentProcessId()
	handle = win32api.OpenProcess(win32con.PROCESS_ALL_ACCESS, True, pid)
	win32process.SetPriorityClass(handle, priorityclasses[priority])

### Check whether it is a possible N-glycan-check
def Func_N_glycan_check (F__description):
	N_glycan_check = True
	F__description = F__description.split (' ')

	Number_of_Hex = 0
	Number_of_HexNAc = 0
	Number_of_dHex = 0
	Number_of_NeuAc = 0
	Number_of_NeuGc = 0

	for i in range(len (F__description)):
		if F__description[i].split('_')[0] == 'Hex':
			Number_of_Hex = int(F__description[i].split('_')[1])
		if F__description[i].split('_')[0] == 'HexNAc':
			Number_of_HexNAc = int(F__description[i].split('_')[1])
		if F__description[i].split('_')[0] == 'dHex':
			Number_of_dHex = int(F__description[i].split('_')[1])
		if F__description[i].split('_')[0] == 'NeuAc':
			Number_of_NeuAc = int(F__description[i].split('_')[1])
		if F__description[i].split('_')[0] == 'NeuGc':
			Number_of_NeuGc = int(F__description[i].split('_')[1])
		Number_of_Sialic = Number_of_NeuAc + Number_of_NeuGc

	if Number_of_Hex >0 and Number_of_HexNAc < 2:
		N_glycan_check = False
	if Number_of_HexNAc >=4 and Number_of_HexNAc < 2:
		N_glycan_check = False
	if Number_of_HexNAc < Number_of_dHex:
		N_glycan_check = False
	if Number_of_dHex >=2 and Number_of_Hex < 2:
		N_glycan_check = False
	if Number_of_dHex >=2 and Number_of_HexNAc < 3:
		N_glycan_check = False
	if Number_of_Sialic ==1 and Number_of_Hex < 3:
		N_glycan_check = False
	if Number_of_Sialic ==1 and Number_of_HexNAc < 3:
		N_glycan_check = False
	if Number_of_Sialic >1 and Number_of_Hex < 4:
		N_glycan_check = False
	if Number_of_Sialic >1 and Number_of_HexNAc < 4:
		N_glycan_check = False
	if Number_of_Hex < Number_of_Sialic:
		N_glycan_check = False

	return N_glycan_check

### generate all possible combinations of precursors or fragments
def Func_combinations (ingredients, mass_max = 6000, charge_max = 1, N_glycan_check = 0):
	F__ranges = [range(i.occur_min, i.occur_max + 1) for i in ingredients]
	F__combinations = []

	for F__counts in itertools.product (*F__ranges):
		F__mass_sum = 0
		F__count_sum = 0
		F__description_parts = []
		F__composition_parts = []

        ### descripition 'Hex_3 HexNAc_2 NeuAc_0 dHex_0'
        ### composition '3 2 0 0'
		for i,j in enumerate (ingredients):
			F__count = F__counts[i]
			F__mass_sum += F__count * j.mass

            ### total number of glycans
			if 'RED' not in j.name and 'ADD' not in j.name and 'H2O' not in j.name:
				F__count_sum += F__count
				F__composition_parts.append ('%d'% (F__count))

			if F__count != 0:
				F__description_parts.append ('%s_%d'% (j.name, F__count))

		F__description = ' '.join (F__description_parts)
		F__composition = '_'.join (F__composition_parts)

        ### add different charges
		if F__mass_sum < (mass_max + 5) and N_glycan_check == 0:
			for i in range(charge_max):
				F__combinations.append (combination (F__description, F__composition, F__mass_sum, i + 1, F__count_sum))

		if F__mass_sum < (mass_max + 5) and N_glycan_check == 1:
			for i in range(charge_max): 
				if Func_N_glycan_check (F__description):
					F__combinations.append (combination (F__description, F__composition, F__mass_sum, i + 1, F__count_sum))

	return F__combinations

### index database
def Func_index_database (input_list, Charged_mass = 1.0073):
	F__output = []

	for i in range (20000):
		F__output.append ([])

	for x in input_list:
		x.mass = x.mass / float (x.charge) + Charged_mass
		F__output[int (x.mass)].append (x)

	return F__output

### filter MGF by spectra list, SN, top-x-peaks
def Func_filter_MGF (MGF_filename, SN, top_x_peaks, spectra_list = 'all'):
	F__SN = SN
	F__top_x_peaks = top_x_peaks
	F__output_MGF = []
	F__spectra_no = 0
	F__spectra_no_output = 0
	F__charge = 1
	F__spectra_no_total = 0

	F__MGF = open (MGF_filename, 'r')
	for line in F__MGF:
		line = line.strip()
		if 'BEGIN' in line:
			F__spectra_no_total += 1

	F__MGF = open (MGF_filename, 'r')
	for line in F__MGF:
		line = line.strip()

        ### start a spectrum with 'BEGIN'
		if 'BEGIN' in line:
			F__charge = 1
			F__peak_mz = []
			F__peak_int = []
			F__peak_int_total = 0
			F__seak_int_avg = 0
			F__spectra_no += 1

            ### print spectral filtering progress
			if F__spectra_no % 1000 == 0 or F__spectra_no == F__spectra_no_total:
				print ('Processing spectra: # Spectra', F__spectra_no, '/', F__spectra_no_total)

        ### spectral general information
		elif 'TITLE' in line:
			F__title = str (line.split ('=')[1])
			F__title = F__title.replace(':','_')
			F__title = F__title.replace('"','_')

		elif 'CHARGE' in line:
			F__charge = int (line.split ('=')[1][0])

		elif 'PEPMASS' in line:
			F__pepmass = float ('%.4f'% float (line.split('=')[1].split(' ')[0]))

		elif 'RTINSECONDS' in line:
			F__rt = float ('%.3f'% float (line.split('=')[1]))

        ### character between m/z and intensity is '\t' in Agilent MGF and ' ' in most other MGF
		elif 'END' not in line and '=' not in line:

			if '\t' in line:
				F__peak_mz_now = float ('%.4f'% float (line.split('\t')[0]))
				F__peak_int_now = float ('%.3f'%float (line.split('\t')[1]))

			elif ' ' in line:
				F__peak_mz_now = float ('%.4f'%float (line.split(' ')[0]))
				F__peak_int_now = float ('%.3f'%float (line.split(' ')[1]))

			F__peak_mz.append (F__peak_mz_now)
			F__peak_int.append (F__peak_int_now)
			F__peak_int_total += F__peak_int_now

        ### output MS/MS spectrum
		elif 'END' in line and ((F__title in spectra_list) or (spectra_list == 'all')):
			F__output_MGF.append ('BEGIN IONS' + '\n')
			F__output_MGF.append ('TITLE=' + str (F__title) + '\n')
			F__output_MGF.append ('CHARGE=' + str (F__charge) + ' + \n')
			F__output_MGF.append ('RTINSECONDS=' + str (F__rt) + '\n')
			F__output_MGF.append ('PEPMASS=' + str (F__pepmass) + '\n')
			F__precursor = F__pepmass * F__charge - (F__charge - 1) * 1.0073
			F__spectra_no_output += 1
			F__peaks_no = len (F__peak_int)
			F__peak_int_sort = sorted (F__peak_int)

            ### peakThreshold1 is based on S/N
			F__peak_threshold1 = F__peak_int_total / F__peaks_no * float (F__SN)

            ### peakThreshold2 is based on number of peaks
			F__peak_threshold2 = 0

			if F__peaks_no > F__top_x_peaks:
				F__peak_threshold2 = F__peak_int_sort[-(F__top_x_peaks + 1)]
				F__peak_int_total = 0

				for peak_intensity in F__peak_int:
					if peak_intensity > F__peak_threshold2:
						F__peak_int_total += peak_intensity

				F__peak_threshold1 = F__peak_int_total / F__top_x_peaks * float (F__SN)

            ### output peak according to PeakThreshold1 and PeakThreshold2
			for i in range(F__peaks_no):
				if F__peak_int[i] > F__peak_threshold1 and F__peak_int[i] > F__peak_threshold2 and F__peak_mz[i] < (F__precursor + 1):
					F__output_MGF.append (str (F__peak_mz[i]) + ' ' + str (F__peak_int[i]) + '\n')

			F__output_MGF.append ('END IONS' + '\n')   

	print ('Total number of outputed spectra =', F__spectra_no_output, '\n')
	return F__output_MGF

### check mass accuracy
def Func_mass_accuracy (precursor, mass_additional, mass_compare, tolerance, tolerance_unit):
	F__output = False

	if tolerance_unit == 'ppm' and abs(float (precursor) - float (mass_additional) - float (mass_compare)) / float (precursor) * 1000000 < float (tolerance):
		F__output = True

	if tolerance_unit == 'Da' and abs(float (precursor) - float (mass_additional) - float (mass_compare)) < float (tolerance):
		F__output = True

	return F__output

### is input a number
def Func_is_number (input):
	try:
		float (input)
		return True
	except ValueError:
		return False

### composition comparison, is 'b' a sub-unit of 'a'?
def Func_composition_inclusion (a,b):
	F__composition_a = a.split ('_')
	F__composition_b = b.split ('_')
	F__output = True

	for i in range (len (F__composition_a)):
		if int(F__composition_a[i]) < int(F__composition_b[i]):
			F__output = False

	return F__output

### (target) divide mass ranges
def Func_mass_ranges (input, min_spe, max_group):

	### group numbers
	F__spe_no = float (len (input))
	if F__spe_no / min_spe < max_group:
		F__group_no = int (F__spe_no / min_spe) + 1
	else:
		F__group_no = max_group

    ### arrange mass list
	F__mass_list = [] 
	for i in range (int (F__spe_no)):
		F__mass_list.append (float('%.4f'% (float (input[i].split ('\t')[0]))))
	F__mass_list.sort()

    ### calculate mass ranges of different groups
	F__output = []
	for i in range (F__group_no):
		F__output.append (F__mass_list[int ((F__spe_no / F__group_no) * i)])
	F__output.append (F__mass_list[-1] + 1)
	return F__output 

### (decoy) add random mass
def Func_add_random (input, range = 1.0):
	F__input = float (input)
	F__range = float (range)
	### range - both
	F__output = random.uniform (F__input * (1 - F__range / 2), F__input * (1 + F__range / 2))
	### range - minus
	#F__output = random.uniform (F__input * (1 - F__range / 2), F__input * (1 + 0))
	### range - plus
	#F__output = random.uniform (F__input * (1 - 0), F__input * (1 + F__range / 2))
	while abs (F__output - F__input) < 0.1 or abs (abs (F__output - F__input) - 1) < 0.1 or abs (abs (F__output - F__input) - 2) < 0.1:
		F__output = random.uniform (F__input * (1 - F__range / 2), F__input * (1 + F__range / 2))
		#F__output = random.uniform (F__input * (1 - F__range / 2), F__input * (1 + 0))
		#F__output = random.uniform (F__input * (1 - 0), F__input * (1 + F__range / 2))
	return F__output

### (decoy) precursor mass based range
def Func_mass_based_range (precursor_mass):
	F__output = 1.0
	if precursor_mass > 2000:
		F__output *= ((2000.0 / precursor_mass) ** 2)
	return F__output

### (decoy) check a precursor against decoy result list, needs more decoy calculation?
def Func_check_decoy (precursor, mass_range, decoy_list, decoy_round):
	F__output = False
	for i in range (len (mass_range) -1):
		F__mass_min = mass_range[i]
		F__mass_max = mass_range[i + 1]
		if F__mass_min <= precursor <= F__mass_max and len (decoy_list[i]) < decoy_round:
			F__output = True
	return F__output

### (decoy) add random calculation to decoy result list
def Func_add_decoy (precursor, mass_range, decoy_list):
	F__output = decoy_list
	for i in range (len (mass_range) -1):
		F__mass_min = mass_range[i]
		F__mass_max = mass_range[i + 1]
		if F__mass_min <= precursor <= F__mass_max:
			F__output[i].append (1)
	return F__output

### (FDR) score distribution
def Func_score_distribution (data, mass_range, distribution_list):
	F__output = []
	for i in range (len (mass_range) -1):
		F__mass_min = mass_range[i]
		F__mass_max = mass_range[i + 1]
		F__total_spe = 0
		for j in range (len (distribution_list)):
			distribution_list[j][0] = 0
		for x in data:
			F__mass_now = float (x.split ('\t')[0])
			F__score_now = float (x.split ('\t')[1])
			if F__mass_min<= F__mass_now <= F__mass_max:
				for j in range (len (distribution_list)):
					if distribution_list[j][1] < F__score_now * 100 <= distribution_list[j][2]:
						distribution_list[j][0] += 1
						F__total_spe += 1
		if F__total_spe == 0:
			F__total_spe = 1
		for j in range(len (distribution_list)):
			F__output.append (str (float (distribution_list[j][0]) / F__total_spe) + '\n')
	return F__output

### (FDR) filter
def Func_FDR_check (precursor, score, FDR_list):
	F__output = True
	for i in range(len (FDR_list)):
		if FDR_list[i][0] <= precursor <= FDR_list[i][1] and score <= FDR_list[i][2]:
			F__output = False
	return F__output

### (FDR) report
def Func_FDR_report (target_data, FDR_final):
	F__output = FDR_final
	for i in range (len (F__output)):
		F__output[i].append (0)
		F__output[i].append (0)
	for x in target_data:
		F__mass_now = float (x.split ('\t')[0])
		F__score_now = float (x.split ('\t')[1])
		for i in range (len (F__output)):
			if F__output[i][0] <= F__mass_now <= F__output[i][1]:
				F__output[i][4] += 1
				if F__score_now > F__output[i][2]:
					F__output[i][3] += 1
	for i in range (len (F__output)):
		F__output[i].append ('%.3f'% (float (F__output[i][3])/float (F__output[i][4])))
	return F__output        

### check spectrum has glycan composition in result
def Func_check_glycan (spectrum, glycan, result):
	F__output = False
	for i in range (len (result)):
		if spectrum == result[i].split('\t')[0] and glycan in result[i].split('\t')[1]:
			F__output = True
	return F__output

###
### read parameters from txt file

OS_dirname, OS_filename = os.path.split(os.path.abspath(Para_filename))
os.chdir (OS_dirname)

Parameter_dirname, Parameter_filename = os.path.split(os.path.abspath(Parameter_file))

File_array_parameter = [line.strip() for line in open (Parameter_file).read().splitlines()]

### thread priority of program
Func_setpriority (None, 0)

### appends names
for i in range (0, len (File_array_parameter), 2): 
	Para_name_array.append (File_array_parameter[i])

### appends values
for i in range (1, len (File_array_parameter), 2): 
	Para_var_array.append (File_array_parameter[i])

### append additionals and margin of error
for i,j in enumerate (Para_name_array):

	if 'Parameter' not in j:

        # get the name and the number
		getvar = j.split (',')

        # the fourth parameters indicates to diagnostic ions
		if len (getvar) == 4:
			MS2_diag_desc.append (getvar[0])
			MS2_diag_mass.append (float (Para_var_array[i]) - 18.010563)
			MS2_diag_mass_orig.append (float (Para_var_array[i]))
			MS2_diag_exist.append (0)
			MS2_diag_exist_orig.append (0)

        # append the proper name back into namearray
		if int (getvar[2]) >0:
			if getvar[0] == 'H2O':
				MS2_fragment_array.append (ingredient (getvar[0], float (Para_var_array[i]), -1, int (getvar[2])))
			else:
				MS2_fragment_array.append (ingredient (getvar[0], float (Para_var_array[i]), 0 , int (getvar[2])))
			MS2_ingredient_array.append (ingredient (getvar[0], float (Para_var_array[i]), int (getvar[1]), int (getvar[2])))

			if 'RED' not in getvar[0] and 'ADD' not in getvar[0] and 'H2O' not in getvar[0]:
				Para_composition_glycan = str (Para_composition_glycan) + str (getvar[0]) + ' '

# append other parameters
	elif 'tolerance_precursor' in j:
		Para_tolerance_precursor = float (Para_var_array[i].split (',')[0])
		Para_tolerance_precursor_unit = str (Para_var_array[i].split (',')[1])

	elif 'tolerance_fragment' in j:
		Para_tolerance_fragment = float (Para_var_array[i].split (',')[0])
		Para_tolerance_fragment_unit = str (Para_var_array[i].split (',')[1])

	elif 'charge_carrier' in j:
		Para_charged_mass = float (Para_var_array[i])

	elif 'see_false' in j:
		Para_temfalse = Para_var_array[i]

	elif 'c13_precursor_analysis' in j:
		Para_c13 = Para_var_array[i]

	elif 'N-glycan filter' in j:
		Para_NGlycan = Para_var_array[i]

    #elif 'filename' in j:
        #Para_filename = Para_var_array[i]

	elif 'file_filter_average_intensity' in j:
		Para_filter_sn = float (Para_var_array[i])

	elif 'file_filter_top_x_peaks' in j:
		Para_filter_top = int (Para_var_array[i])

	elif 'file_filter_diagnostic_ions' in j:
		Para_filter_diagnostic_ions = Para_var_array[i]

### after getting all user decided masses, create all combinations   
MS2_precursor_list = Func_index_database (Func_combinations (MS2_ingredient_array, MS2_precursor_mass_max, 1, Para_NGlycan), Para_charged_mass)
MS2_fragment_list = Func_index_database (Func_combinations (MS2_fragment_array, MS2_precursor_mass_max, Test_charge_max), Para_charged_mass)

### c13 precursor analysis
c13_precursor_shift = 1.0033
if Para_c13 == 'no':
	c13_precursor_shift = 9999

### N-glycan filtration
if Para_NGlycan == 'yes':
	Para_NGlycan = 1
else:
	Para_NGlycan = 0

###
### main program (1st step - spectral filtration)

print ('\n1st step - spectral filtration\n(Diectory)', OS_dirname, '\n(Filename)', OS_filename, '\n(Parameter)', Parameter_filename, '\n')

### filter the original spectra by S/N ration and peak number, and output standard MGF format

### filter without diagnostic ions
if Para_filter_diagnostic_ions != 'yes':
	MGF_filtered_final = Func_filter_MGF (Para_filename, Para_filter_sn, Para_filter_top)

### filter with diagnostic ions
if Para_filter_diagnostic_ions == 'yes':
	MGF_filtered_diagnostic_ions = Func_filter_MGF (Para_filename, 0, Para_filter_diagnostic_ions_top)
	File_array = [line.strip() for line in MGF_filtered_diagnostic_ions]

	print ('...filtering spectra using diagnostic ions...\n')
	Spectra_no_filtering = 0

	MS2_fragment_list_diagnostic = []
	for i in range (len (MS2_fragment_list)):
		for x in MS2_fragment_list[i]:
			if x.count <= Para_filter_diagnostic_ions_count and x.count >= 1:
				MS2_fragment_list_diagnostic.append (x)

	j = 0
	while j < len (File_array):
		if 'BEGIN' in File_array[j]:
			MS2_title = str (File_array[j + 1].split('=')[1])
			MS2_charge = int (File_array[j + 2].split ('=')[1][0])
			MS2_pepmass = float (File_array[j + 4].split('=')[1])
			MS2_precursor = MS2_pepmass * float (MS2_charge) - (MS2_charge - 1) * Para_charged_mass
			j += 5

			MS2_diagnostic_check = False

            ### check diagnostic ions in diagnostic lists or neutral loss from precursor
			while 'IONS' not in File_array[j] and '=' not in File_array[j]:
				MS2_peak_now = File_array[j].split(' ')
				MS2_peak_now_mz = float ('%.4f'% float (MS2_peak_now[0]))

				if MS2_diagnostic_check == False:
					for x in MS2_fragment_list_diagnostic:
						if Func_mass_accuracy (MS2_peak_now_mz, 0, x.mass, Para_tolerance_fragment, Para_tolerance_fragment_unit) and x.charge == 1:
							MS2_diagnostic_check = True
						else:
							for i in range (1, MS2_charge + 1):
								if Func_mass_accuracy (MS2_precursor, (MS2_peak_now_mz * i - (i - 1) * Para_charged_mass), x.mass, Para_tolerance_fragment, Para_tolerance_fragment_unit) and x.count == 1:
									MS2_diagnostic_check = True
				j += 1

        ### list of MS/MS spectra with diagnostic ions
		if 'END' in File_array[j]:
			Spectra_no_filtering += 1
			if MS2_diagnostic_check:
				MGF_list_diagnostic_ions.append (MS2_title)
		if Spectra_no_filtering % 1000 == 0:
			print ('Filtering spectra: # Spectra', Spectra_no_filtering)
		j += 1

	print ('\n...output filtered spectra with diagnostic ions...\n')
    # output MGF with diagnostic ions
	MGF_filtered_final = Func_filter_MGF (Para_filename, Para_filter_sn, Para_filter_top, MGF_list_diagnostic_ions)

### read the filtered MGF
File_array = [line.strip() for line in MGF_filtered_final]

### the number of spectra and max precursor mass in current file
MS2_precursor_mass_max = 0

### get the maximum precursor mass
for j in range (len (File_array)):
	if 'BEGIN' in File_array[j]:
		MS2_spectra_no += 1
		MS2_charge = int (File_array[j + 2].split ('=')[1][0])
		MS2_pepmass = float (File_array[j + 4].split('=')[1])
		MS2_precursor = MS2_pepmass * float (MS2_charge) - (MS2_charge-1) * Para_charged_mass
		if MS2_precursor > MS2_precursor_mass_max:
			MS2_precursor_mass_max = MS2_precursor

###
### main program (2nd step - target)

print ('\n2nd step - target\n(Diectory)', OS_dirname, '\n(Filename)', OS_filename, '\n(Parameter)', Parameter_filename, '\n')

Step2_time_start = time.time()

for i in range (MS2_spectra_no):
	File_result_target_int.append ('0\n')
	File_result_target_seq.append ('0\n')

MS2_diag_desc = []
MS2_diag_mass = []
MS2_diag_mass_orig = []
MS2_diag_exist = []
MS2_diag_exist_orig = []

MS2_fragment_array = []
MS2_ingredient_array = []

### get user-defined number of monosaccharides
for i,j in enumerate (Para_name_array):

	if 'Parameter' not in j:

        # get the name and the number
		getvar = j.split (',')

        # the fourth parameters indicates to diagnostic ions
		if len (getvar) == 4:
			MS2_diag_desc.append (getvar[0])
			MS2_diag_mass.append (float (Para_var_array[i]) - 18.010563)
			MS2_diag_mass_orig.append (float (Para_var_array[i]))
			MS2_diag_exist.append (0)
			MS2_diag_exist_orig.append (0)

        # append the proper name back into namearray
		if int (getvar[2]) >0:
			if getvar[0] == 'H2O':
				MS2_fragment_array.append (ingredient (getvar[0], float (Para_var_array[i]), -1, int (getvar[2])))
			else:
				MS2_fragment_array.append (ingredient (getvar[0], float (Para_var_array[i]), 0 , int (getvar[2])))
			MS2_ingredient_array.append (ingredient (getvar[0], float (Para_var_array[i]), int (getvar[1]), int (getvar[2])))

### after getting all user decided masses, create all combinations   
MS2_precursor_list = Func_index_database (Func_combinations (MS2_ingredient_array, MS2_precursor_mass_max, 1, Para_NGlycan), Para_charged_mass)
MS2_fragment_list = Func_index_database (Func_combinations (MS2_fragment_array, MS2_precursor_mass_max, Test_charge_max), Para_charged_mass)

MS2_spectra_no_now = 0

### calculation of target database
j = 0
while j < len (File_array):

	if 'BEGIN' in File_array[j]:
		MS2_spectra_no_now += 1
		if MS2_spectra_no_now % 1000 == 0 or MS2_spectra_no_now == MS2_spectra_no:
			print ('Processing spectra: # ', MS2_spectra_no_now, ' / total', MS2_spectra_no)

        ### reset the diagnostic ions
		for i in range (len (MS2_diag_exist)):
			MS2_diag_exist[i] = 0
			MS2_diag_exist_orig[i] = 0

        ### spectral general information
		MS2_charge = int (File_array[j + 2].split ('=')[1][0])
		MS2_pepmass = float ('%.4f' % (float (File_array[j + 4].split('=')[1])))
		MS2_precursor = MS2_pepmass * float (MS2_charge) - (MS2_charge-1) * Para_charged_mass

		j += 5

        ### store original array position incase we need to look at the same spectrum multiple times
		j_indexstore = j

		while 'IONS' not in File_array[j] and '=' not in File_array[j]:
			MS2_peak_now = File_array[j].split (' ')

            ### generate the diagnostic information
			for i in range(len(MS2_diag_exist)):
				if Func_mass_accuracy (MS2_peak_now[0], Para_charged_mass, MS2_diag_mass[i], Para_tolerance_fragment, Para_tolerance_fragment_unit):
					MS2_diag_exist[i] = 1
				if Func_mass_accuracy (MS2_peak_now[0], Para_charged_mass, MS2_diag_mass_orig[i], Para_tolerance_fragment, Para_tolerance_fragment_unit):
					MS2_diag_exist_orig[i] = 1
			j += 1

        ### generate combinations for every precursor value
		for x in (MS2_precursor_list[int (MS2_precursor)-1] + MS2_precursor_list[int (MS2_precursor)] + MS2_precursor_list[int (MS2_precursor) + 1]):
			MS2_diag_check = True

        ### check the diagnostic information
			for i in range (len (MS2_diag_exist)):
				if MS2_diag_exist[i] == 0 and MS2_diag_desc[i] in x.description:
					MS2_diag_check = False
				if MS2_diag_exist[i] == 1 and MS2_diag_exist_orig[i] == 1 and MS2_diag_desc[i] not in x.description:
					MS2_diag_check = False

            ### find precursor combination
			if (Func_mass_accuracy (MS2_precursor, 0, x.mass, Para_tolerance_precursor, Para_tolerance_precursor_unit) or Func_mass_accuracy (MS2_precursor, c13_precursor_shift, x.mass, Para_tolerance_precursor, Para_tolerance_precursor_unit)) and MS2_diag_check and x.count > 1:
				MS2_precursor_composition = x.composition
				MS2_precursor_description = x.description
				j = j_indexstore
				MS2_c13 = 0

				MS2_seq_cov_data = []
				for i in range (x.count):
					MS2_seq_cov_data.append (False)

                ### matches C13 peak
				if Func_mass_accuracy (MS2_precursor, c13_precursor_shift, x.mass, Para_tolerance_precursor, Para_tolerance_precursor_unit):
					MS2_precursor -= c13_precursor_shift
					MS2_c13 = 1 

                ### set a variable to calculate total abundance for this run
				MS2_intensity_sum = 0.0001
				MS2_intensity_sum_all = 0.0
				MS2_intensity_sum_real = 0.0
				MS2_intensity_sum_false = 0.0
				MS2_intensity_max = 0.0
				MS2_cov_int = 0.0
				MS2_cov_seq = 0.0

				MS2_real_mass_exp = []
				MS2_real_abundance = []
				MS2_real_mass_theo = []
				MS2_real_desc = []
				MS2_real_comp = []

				while 'IONS' not in File_array[j] and '=' not in File_array[j]:
					MS2_peak_now = File_array[j].split (' ')
					MS2_peak_now_mz = float (MS2_peak_now[0])
					MS2_peak_now_inten = float (MS2_peak_now[1])

                    ### total intensity for MS/MS ions
					MS2_intensity_sum_all += MS2_peak_now_inten

                    ### first match peaks with real fragments
					for i in range(1, Test_charge_max + 1):
						for y in (MS2_fragment_list[int (MS2_peak_now_mz)-1] + MS2_fragment_list[int (MS2_peak_now_mz)] + MS2_fragment_list[int (MS2_peak_now_mz) + 1]):
							if y.charge == i and y.charge <= MS2_charge and Func_mass_accuracy (MS2_peak_now_mz, 0, y.mass, Para_tolerance_fragment, Para_tolerance_fragment_unit) and MS2_peak_now_inten > 0:
								MS2_fragment_composition = y.composition
								if Func_composition_inclusion (MS2_precursor_composition, MS2_fragment_composition):
									MS2_intensity_sum_real += MS2_peak_now_inten
									MS2_peak_now_inten = 0
									MS2_seq_cov_data[y.count - 1] = True

                    ### this advances iterator inside each test chunk
					j += 1

				MS2_intensity_sum += MS2_intensity_sum_real

				if MS2_c13 == 1:
					MS2_precursor += c13_precursor_shift

				MS2_seq_cov = 0.0
				for i in range (x.count - 1):
					if MS2_seq_cov_data[i] == True:
						MS2_seq_cov += 1

                ### output result for current glycan candidate
				if File_result_target_int[MS2_spectra_no_now - 1] == ('0\n'):
					File_result_target_int[MS2_spectra_no_now - 1] = (str (MS2_precursor) + '\t' + str ('%.3f'% (float (MS2_intensity_sum_real/MS2_intensity_sum_all))) + '\n')
				elif float (File_result_target_int[MS2_spectra_no_now - 1].split('\t')[1]) < float (MS2_intensity_sum_real/MS2_intensity_sum_all):
					File_result_target_int[MS2_spectra_no_now - 1] = (str (MS2_precursor) + '\t' + ('%.3f'% (float (MS2_intensity_sum_real/MS2_intensity_sum_all))) + '\n')

                ### output result for current glycan candidate
				if File_result_target_seq[MS2_spectra_no_now - 1] == ('0\n'):
					File_result_target_seq[MS2_spectra_no_now - 1] = (str (MS2_precursor) + '\t' + str ('%.3f'% (float (MS2_seq_cov / (x.count - 1)))) + '\n')
				elif float (File_result_target_seq[MS2_spectra_no_now - 1].split('\t')[1]) < float (MS2_seq_cov / (x.count - 1)):
					File_result_target_seq[MS2_spectra_no_now - 1] = (str (MS2_precursor) + '\t' + ('%.3f'% (float (MS2_seq_cov / (x.count - 1)))) + '\n')

	j += 1

FDR_target_data_int = []
FDR_target_data_seq = []

### target data (score > 0)
for i in range (len (File_result_target_int)):
	if '\t' in File_result_target_int[i]:
		FDR_target_data_int.append (File_result_target_int[i])
		FDR_target_data_seq.append (File_result_target_seq[i])

### mass ranges for target data            
FDR_mass_ranges = Func_mass_ranges (FDR_target_data_int, Test_min_spe, Test_max_group)

Step2_time_end = time.time()

print ('Total time for target analysis = ', ('--- %.0fs seconds' % (Step2_time_end - Step2_time_start)))
print ('Maximum time for decoy analysis = ', ('--- %.0fs seconds' % ((Step2_time_end - Step2_time_start) * Decoy_vs_target_repeat)), ('(--- %.0fs) seconds' % (Decoy_repeat_max_time)))

###
### main program (3rd step - decoy)

print ('\n3rd step - decoy\n(Diectory)', OS_dirname, '\n(Filename)', OS_filename, '\n(Parameter)', Parameter_filename, '\n')

FDR_decoy_data_int = []
FDR_decoy_data_seq = []
Print_decoy_progress = []

for i in range (len (FDR_mass_ranges) - 1):
	File_result_decoy.append ([])
	Print_decoy_progress.append (0)

Time_start_decoy = time.time()
Para_decoy_progress = True
Decoy_round_total = int (float (len (FDR_target_data_int)) / float (len (FDR_mass_ranges) - 1) * Para_decoy_round)
Decoy_repeat = 0

### calculation of decoy databases
while Para_decoy_progress:

    ### create random database
	MS2_diag_desc = []
	MS2_diag_mass = []
	MS2_diag_mass_orig = []
	MS2_diag_exist = []
	MS2_diag_exist_orig = []
	MS2_fragment_array = []
	MS2_ingredient_array = []

	for i,j in enumerate (Para_name_array):
		if 'Parameter' not in j:
            ### get the name and the number
			getvar = j.split (',')

            ### the fourth parameters indicates to diagnostic ions
			if len (getvar) == 4:
				MS2_diag_desc.append (getvar[0])
				MS2_diag_mass.append (float (Para_var_array[i]) - 18.010563)
				MS2_diag_mass_orig.append (float (Para_var_array[i]))
				MS2_diag_exist.append (0)
				MS2_diag_exist_orig.append (0)

			if int (getvar[2]) >0:
				if getvar[0] == 'H2O':
					MS2_fragment_array.append (ingredient (getvar[0], float (Para_var_array[i]), -1, int (getvar[2])))
					MS2_ingredient_array.append (ingredient (getvar[0], float (Para_var_array[i]), int (getvar[1]), int (getvar[2])))
				elif ('Neu' not in getvar[0]) and ('ADD' not in getvar[0]) and ('RED' not in getvar[0]) and ('SiaNAz' not in getvar[0]):
					MS2_fragment_array.append (ingredient (getvar[0], Func_add_random (Para_var_array[i], Test_random_range), 0, int (getvar[2])))
					MS2_ingredient_array.append (ingredient (getvar[0], Func_add_random (Para_var_array[i], Test_random_range), int (getvar[1]), int (getvar[2])))
				else:
					MS2_fragment_array.append (ingredient (getvar[0], float (Para_var_array[i]), 0, int (getvar[2])))
					MS2_ingredient_array.append (ingredient (getvar[0], float (Para_var_array[i]), int (getvar[1]), int (getvar[2])))

    ### after getting all user decided masses, create all combinations   
	MS2_precursor_list = Func_index_database (Func_combinations (MS2_ingredient_array, MS2_precursor_mass_max, 1, Para_NGlycan), Para_charged_mass)
	MS2_fragment_list = Func_index_database (Func_combinations (MS2_fragment_array, MS2_precursor_mass_max, Test_charge_max), Para_charged_mass)

	j = 0
	while j < len (File_array):

		if 'BEGIN' in File_array[j]:

            ### reset the diagnostic ions
			for i in range (len (MS2_diag_exist)):
				MS2_diag_exist[i] = 0
				MS2_diag_exist_orig[i] = 0

            ### spectral general information
			MS2_charge = int (File_array[j + 2].split ('=')[1][0])
			MS2_pepmass = float ('%.4f' % (float (File_array[j + 4].split('=')[1])))
			MS2_precursor = MS2_pepmass * float (MS2_charge) - (MS2_charge-1) * Para_charged_mass

			j += 5

            ### store original array position incase we need to look at the same spectrum multiple times
			j_indexstore = j

			while 'IONS' not in File_array[j] and '=' not in File_array[j]:
				MS2_peak_now = File_array[j].split (' ')

                ### generate the diagnostic information
				for i in range (len (MS2_diag_exist)):
					if Func_mass_accuracy (MS2_peak_now[0], Para_charged_mass, MS2_diag_mass[i], Para_tolerance_fragment, Para_tolerance_fragment_unit):
						MS2_diag_exist[i] = 1
					if Func_mass_accuracy (MS2_peak_now[0], Para_charged_mass, MS2_diag_mass_orig[i], Para_tolerance_fragment, Para_tolerance_fragment_unit):
						MS2_diag_exist_orig[i] = 1
				j += 1

			if Func_check_decoy (MS2_precursor, FDR_mass_ranges, File_result_decoy, Decoy_round_total):
                ### generate combinations for every precursor value
				for x in (MS2_precursor_list[int (MS2_precursor)-1] + MS2_precursor_list[int (MS2_precursor)] + MS2_precursor_list[int (MS2_precursor) + 1]):
					MS2_diag_check = True

                    ### check the diagnostic information
					for i in range (len (MS2_diag_exist)):
						if MS2_diag_exist[i] == 0 and MS2_diag_desc[i] in x.description:
							MS2_diag_check = False
						if MS2_diag_exist[i] == 1 and MS2_diag_exist_orig[i] == 1 and MS2_diag_desc[i] not in x.description:
							MS2_diag_check = False

                    ### find precursor combination
					if (Func_mass_accuracy (MS2_precursor, 0, x.mass, Func_mass_based_range (MS2_precursor), 'Da') or Func_mass_accuracy (MS2_precursor, c13_precursor_shift, x.mass, Func_mass_based_range (MS2_precursor), 'Da')) and MS2_diag_check and x.count > 1:
						MS2_precursor_composition = x.composition
						MS2_precursor_description = x.description
						j = j_indexstore
						MS2_c13 = 0
						
						MS2_seq_cov_data = []
						for i in range (x.count):
							MS2_seq_cov_data.append (False)

                        ### matches C13 peak
						if Func_mass_accuracy (MS2_precursor, c13_precursor_shift, x.mass, Para_tolerance_precursor, Para_tolerance_precursor_unit):
							MS2_precursor -= c13_precursor_shift
							MS2_c13 = 1 

                        ### set a variable to calculate total abundance for this run
						MS2_intensity_sum = 0.0001
						MS2_intensity_sum_all = 0.0
						MS2_intensity_sum_real = 0.0
						MS2_intensity_sum_false = 0.0
						MS2_intensity_max = 0.0
						MS2_cov_int = 0.0
						MS2_cov_seq = 0.0

						MS2_real_mass_exp = []
						MS2_real_abundance = []
						MS2_real_mass_theo = []
						MS2_real_desc = []
						MS2_real_comp = []

						while 'IONS' not in File_array[j] and '=' not in File_array[j]:
							MS2_peak_now = File_array[j].split (' ')
							MS2_peak_now_mz = float (MS2_peak_now[0])
							MS2_peak_now_inten = float (MS2_peak_now[1])

                            ### total intensity for MS/MS ions
							MS2_intensity_sum_all += MS2_peak_now_inten
                            ### first match peaks with real fragments
							for i in range(1, Test_charge_max + 1):
								for y in (MS2_fragment_list[int (MS2_peak_now_mz)-1] + MS2_fragment_list[int (MS2_peak_now_mz)] + MS2_fragment_list[int (MS2_peak_now_mz) + 1]):
									if y.charge == i and y.charge <= MS2_charge and Func_mass_accuracy (MS2_peak_now_mz, 0, y.mass, Para_tolerance_fragment, Para_tolerance_fragment_unit) and MS2_peak_now_inten > 0:
										MS2_fragment_composition = y.composition
										if Func_composition_inclusion (MS2_precursor_composition, MS2_fragment_composition):
											MS2_intensity_sum_real += MS2_peak_now_inten
											MS2_peak_now_inten = 0
											MS2_seq_cov_data[y.count - 1] = True

                            ### this advances iterator inside each test chunk
							j += 1

						MS2_intensity_sum += MS2_intensity_sum_real
						MS2_intensity_sum += MS2_intensity_sum_false
						
						if MS2_c13 == 1:
							MS2_precursor += c13_precursor_shift

						MS2_seq_cov = 0.0
						for i in range (x.count - 1):
							if MS2_seq_cov_data[i] == True:
								MS2_seq_cov += 1

                        ### record decoy result
						File_result_decoy = Func_add_decoy (MS2_precursor, FDR_mass_ranges, File_result_decoy)

                        ### output result for current glycan candidate
						FDR_decoy_data_int.append (str (MS2_precursor) + '\t' + ('%.3f'% (float (MS2_intensity_sum_real/MS2_intensity_sum_all))) + '\n')
						if x.count > 1:
							FDR_decoy_data_seq.append (str (MS2_precursor) + '\t' + ('%.3f'% (float (MS2_seq_cov / (x.count - 1)))) + '\n')
						else:
							FDR_decoy_data_seq.append (str (MS2_precursor) + '\t0\n')

		j += 1

	Decoy_repeat += 1

	Para_decoy_progress = False
	for i in range (len (FDR_mass_ranges) - 1):
		Print_decoy_progress[i] = (int (len (File_result_decoy[i]) * 1000000 / Decoy_round_total))
		print ('mass range', FDR_mass_ranges[i], '-', FDR_mass_ranges[i + 1], ' : ', Print_decoy_progress[i], '%')
		if len (File_result_decoy[i]) < Decoy_round_total:
			Para_decoy_progress = True

	if Decoy_repeat > Decoy_repeat_max_round:
		Para_decoy_progress = False
	if time.time() - Time_start_decoy > ((Step2_time_end - Step2_time_start) * Decoy_vs_target_repeat) or time.time() - Time_start_decoy > Decoy_repeat_max_time:
		Para_decoy_progress = False

	print ('Decoy_repeat_#', Decoy_repeat, '  Decoy_time_#', ('--- %.0fs' % (time.time() - Time_start_decoy)), '\n', 'Maximum time for decoy analysis = ', ('--- %.0fs seconds' % ((Step2_time_end - Step2_time_start) * Decoy_vs_target_repeat)), ('(--- %.0fs) seconds' % (Decoy_repeat_max_time)), '\n')

###
### main program (4th step - FDR)

print ('\n4th step - FDR\n(Diectory)', OS_dirname, '\n(Filename)', OS_filename, '\n(Parameter)', Parameter_filename, '\n')

### score distribution
FDR_distribution_list = []
for i in range(100):
	FDR_distribution_list.append ([0, 99 - i, 99 - i + 1])

FDR_distribution_list[99][1] -= 1

FDR_decoy_dis_int = Func_score_distribution (FDR_decoy_data_int, FDR_mass_ranges, FDR_distribution_list)
FDR_target_dis_int = Func_score_distribution (FDR_target_data_int, FDR_mass_ranges, FDR_distribution_list)
FDR_decoy_dis_seq = Func_score_distribution (FDR_decoy_data_seq, FDR_mass_ranges, FDR_distribution_list)
FDR_target_dis_seq = Func_score_distribution (FDR_target_data_seq, FDR_mass_ranges, FDR_distribution_list)

### FDR_int
FDR_decoy_array = [line.strip() for line in FDR_decoy_dis_int]
FDR_target_array = [line.strip() for line in FDR_target_dis_int]

FDR_value = []
FDR_decoy_total = []
FDR_target_total = []
FDR_current_position = 0

for i in range (100):
	FDR_decoy_total.append (0.0)
	FDR_target_total.append (0.0)

### calculation of score threshold
i = 0
while i < len (FDR_decoy_array):
	FDR_current_score = i % 100
	for j in range(FDR_current_score + 1):
		FDR_decoy_total[FDR_current_score] += float (FDR_decoy_array[j + FDR_current_position])
		FDR_target_total[FDR_current_score] += float (FDR_target_array[j + FDR_current_position])

	if FDR_current_score % 100 == 99:
		FDR_score_check = True
		FDR_score_now = 99
		while FDR_score_check:
			if FDR_target_total[FDR_score_now] == 0 or FDR_score_now == 0:
				FDR_value.append (FDR_score_now)
				FDR_score_check = False
				FDR_current_position += 100
			elif FDR_decoy_total[FDR_score_now] / FDR_target_total[FDR_score_now] < FDR_threshold_int:
				FDR_value.append (FDR_score_now)
				FDR_score_check = False
				FDR_current_position += 100
			FDR_score_now -= 1
		for j in range (100):
			FDR_decoy_total[j] = 0.0
			FDR_target_total[j] = 0.0
	i += 1

FDR_final_int = []

for i in range (len (FDR_value)):
	FDR_final_int.append ([FDR_mass_ranges[i], FDR_mass_ranges[i + 1], float (99 - FDR_value[i]) / 100])

FDR_report = Func_FDR_report (FDR_target_data_int, FDR_final_int)

### output score distribution of target and decoy, intensity
if Test_output_FDR:
	with open (Para_filename.split ('.')[0] + '-Result-FDR-Int.txt', 'w') as Output_FDR_report:
		Output_FDR_report.write ('Mass_range_1\tMass_range_2\tFDR-Cutoff\t#ID\t#Total\tID/Total\n')
		for i in range(len (FDR_report)):
			for x in FDR_report[i]:
				Output_FDR_report.write ('%s\t' % x)
			Output_FDR_report.write ('\n')

if Test_output_distribution:
	with open (Para_filename.split ('.')[0] + '-Result-dis-int-decoy.txt', 'w') as Output_File_result_decoy_dis:
		for line in FDR_decoy_dis_int:
			Output_File_result_decoy_dis.write (line)

	with open (Para_filename.split ('.')[0] + '-Result-dis-int-target.txt', 'w') as Output_File_result_target_dis:
		for line in FDR_target_dis_int:
			Output_File_result_target_dis.write (line)


print ('\nFDR of intensity coverage')
for i in range (len (FDR_report)):
	print ('mass range(' + str (FDR_report[i][0]) + '-' + str (FDR_report[i][1]) + '):  FDR Cutoff- (' + str ('%.3f'% (FDR_report[i][2])) + ')  #ID/Total- ' + str (FDR_report[i][5]) + ' (' + str (FDR_report[i][3]) + '/' + str (FDR_report[i][4]) + ')')

### FDR_seq
FDR_decoy_array = [line.strip() for line in FDR_decoy_dis_seq]
FDR_target_array = [line.strip() for line in FDR_target_dis_seq]

FDR_value = []
FDR_decoy_total = []
FDR_target_total = []
FDR_current_position = 0

for i in range (100):
	FDR_decoy_total.append (0.0)
	FDR_target_total.append (0.0)

i = 0
while i < len (FDR_decoy_array):
	FDR_current_score = i % 100
	for j in range (FDR_current_score + 1):
		FDR_decoy_total[FDR_current_score] += float (FDR_decoy_array[j + FDR_current_position])
		FDR_target_total[FDR_current_score] += float (FDR_target_array[j + FDR_current_position])

	if FDR_current_score % 100 == 99:
		FDR_score_check = True
		FDR_score_now = 99
		while FDR_score_check:
			if FDR_target_total[FDR_score_now] == 0 or FDR_score_now == 0:
				FDR_value.append (FDR_score_now)
				FDR_score_check = False
				FDR_current_position += 100		  
			elif FDR_decoy_total[FDR_score_now] / FDR_target_total[FDR_score_now] < FDR_threshold_seq:
				FDR_value.append (FDR_score_now)
				FDR_score_check = False
				FDR_current_position += 100
			FDR_score_now -= 1
		for j in range(100):
			FDR_decoy_total[j] = 0.0
			FDR_target_total[j] = 0.0
	i += 1

FDR_final_seq = []

for i in range (len (FDR_value)):
	FDR_final_seq.append ([FDR_mass_ranges[i], FDR_mass_ranges[i + 1], float (99 - FDR_value[i]) / 100])

FDR_report = Func_FDR_report (FDR_target_data_seq, FDR_final_seq)

print ('\nFDR of sequence coverage')
for i in range (len (FDR_report)):
	print ('mass range(' + str (FDR_report[i][0]) + '-' + str (FDR_report[i][1]) + '):  FDR Cutoff- (' + str ('%.3f'% (FDR_report[i][2])) + ')  #ID/Total- ' + str (FDR_report[i][5]) + ' (' + str (FDR_report[i][3]) + '/' + str (FDR_report[i][4]) + ')')

### output decoy list

if Test_output_decoy_list:
	with open(Para_filename.split ('.')[0] + '-Result-decoy-list.txt', 'w') as Output_decoy_list:
		Output_decoy_list.write ('Precursor\tCov.int\tCov.seq\n')
		for i in range (len (FDR_decoy_data_int)):
			Output_decoy_list.write (FDR_decoy_data_int[i].split('\t')[0])
			Output_decoy_list.write ('\t')
			Output_decoy_list.write (FDR_decoy_data_int[i].split('\t')[1].split('\n')[0])
			Output_decoy_list.write ('\t')
			Output_decoy_list.write (FDR_decoy_data_seq[i].split('\t')[1].split('\n')[0])
			Output_decoy_list.write ('\n')

### output score distribution of target and decoy, intensity
if Test_output_FDR:
	with open (Para_filename.split ('.')[0] + '-Result-FDR-Seq.txt', 'w') as Output_FDR_report:
		Output_FDR_report.write ('Mass_range_1\tMass_range_2\tFDR-Cutoff\t#ID\t#Total\tID/Total\n')
		for i in range(len (FDR_report)):
			for x in FDR_report[i]:
				Output_FDR_report.write ('%s\t' % x)
			Output_FDR_report.write ('\n')

if Test_output_distribution:
	with open (Para_filename.split ('.')[0] + '-Result-dis-seq-decoy.txt', 'w') as Output_File_result_decoy_dis:
		for line in FDR_decoy_dis_seq:
			Output_File_result_decoy_dis.write (line)

	with open (Para_filename.split ('.')[0] + '-Result-dis-seq-target.txt', 'w') as Output_File_result_target_dis:
		for line in FDR_target_dis_seq:
			Output_File_result_target_dis.write (line)

###
### main program (5th step - result and annotation)

print ('\n5th step - result and annotation\n(Diectory)', OS_dirname, '\n(Filename)', OS_filename, '\n(Parameter)', Parameter_filename, '\n')

MS2_diag_desc = []
MS2_diag_mass = []
MS2_diag_mass_orig = []
MS2_diag_exist = []
MS2_diag_exist_orig = []

MS2_fragment_array = []
MS2_ingredient_array = []

for i,j in enumerate (Para_name_array):

	if 'Parameter' not in j:

        # get the name and the number
		getvar = j.split (',')

        # the fourth parameters indicates to diagnostic ions
		if len (getvar) == 4:
			MS2_diag_desc.append (getvar[0])
			MS2_diag_mass.append (float (Para_var_array[i]) - 18.010563)
			MS2_diag_mass_orig.append (float (Para_var_array[i]))
			MS2_diag_exist.append (0)
			MS2_diag_exist_orig.append (0)

        # append the proper name back into namearray
		if int (getvar[2]) >0:
			if getvar[0] == 'H2O':
				MS2_fragment_array.append (ingredient (getvar[0], float (Para_var_array[i]), -1, int (getvar[2])))
			else:
				MS2_fragment_array.append (ingredient (getvar[0], float (Para_var_array[i]), 0 , int (getvar[2])))
			MS2_ingredient_array.append (ingredient (getvar[0], float (Para_var_array[i]), int (getvar[1]), int (getvar[2])))

### here N-glycan-check is considered in precursor list
MS2_precursor_list = Func_index_database (Func_combinations (MS2_ingredient_array, MS2_precursor_mass_max, 1, Para_NGlycan), Para_charged_mass)
MS2_fragment_list = Func_index_database (Func_combinations (MS2_fragment_array, MS2_precursor_mass_max, Test_charge_max), Para_charged_mass)

if not os.path.exists (Para_filename.split ('.')[0] + '-annotated-spectra') and Test_output_spe:
	os.mkdir (Para_filename.split ('.')[0] + '-annotated-spectra')

if Test_output_spe:
	os.chdir (Para_filename.split ('.')[0] + '-annotated-spectra')

### result headline
File_result.append ('spectra\tretention\tprecursor\tglycan mass\tPPM\t' + Para_composition_glycan + '\tdescription\tcharge\tCov. Int\tCov. Seq\tC13\tTotal MS/MS intensity\n')

MS2_precursor_description = 'b'
MS2_printed = True

MS2_identified_list = []
MS2_identified_title = []
MS2_identified_current_score = 0

print ('\nprocessing outputed spectrum list')
MS2_spectra_no_now = 0

j = 0
while j < len (File_array):

    ### nonassignable list
	if MS2_precursor_description == 'a' and MS2_printed:
		MS2_nonassignable.append (str (MS2_title) + '\t' + str (MS2_pepmass) + '\t' + str (MS2_rt) + '\tDoes not fit any combination\t\t' + str (MS2_charge) + '\t\n')
		MGF_list_nonassignable.append (MS2_title)
		MS2_printed = False

	if 'BEGIN' in File_array[j]:
        ### progress of spectral processing
		MS2_spectra_no_now += 1
		if (MS2_spectra_no_now % 1000 == 0 or MS2_spectra_no_now == MS2_spectra_no):
			print ('Processing spectra: # ', MS2_spectra_no_now, ' / total', MS2_spectra_no)

        ### reset the diagnostic ions
		for i in range (len (MS2_diag_exist)):
			MS2_diag_exist[i] = 0
			MS2_diag_exist_orig[i] = 0

        ### spectral general information
		MS2_title = str (File_array[j + 1].split ('=')[1])
		MS2_charge = int (File_array[j + 2].split ('=')[1][0])
		MS2_rt = float ('%.3f' % ((float (File_array[j + 3].split('=')[1]))/60))
		MS2_pepmass = float ('%.4f' % (float (File_array[j + 4].split('=')[1])))
		MS2_precursor = MS2_pepmass * float (MS2_charge) - (MS2_charge-1) * Para_charged_mass
		MS2_precursor_description = 'a'
		MS2_printed = True

		j += 5

        ### store original array position incase we need to look at the same peptide multiple times
		j_indexstore = j

		while 'IONS' not in File_array[j] and '=' not in File_array[j]:
			MS2_peak_now = File_array[j].split (' ')

            ### generate the diagnostic information
			for i in range (len (MS2_diag_exist)):
				if Func_mass_accuracy (MS2_peak_now[0], Para_charged_mass, MS2_diag_mass[i], Para_tolerance_fragment, Para_tolerance_fragment_unit):
					MS2_diag_exist[i] = 1
				if Func_mass_accuracy (MS2_peak_now[0], Para_charged_mass, MS2_diag_mass_orig[i], Para_tolerance_fragment, Para_tolerance_fragment_unit):
					MS2_diag_exist_orig[i] = 1
			j += 1

        ### generate combinations for every precursor value
		for x in (MS2_precursor_list[int (MS2_precursor)-1] + MS2_precursor_list[int (MS2_precursor)] + MS2_precursor_list[int (MS2_precursor) + 1]):
			MS2_diag_check = True

        ### check the diagnostic information
			for i in range (len (MS2_diag_exist)):
				if MS2_diag_exist[i] == 0 and MS2_diag_desc[i] in x.description:
					MS2_diag_check = False
				if MS2_diag_exist[i] == 1 and MS2_diag_exist_orig[i] == 1 and MS2_diag_desc[i] not in x.description:
					MS2_diag_check = False

            ### find precursor combination
			if (Func_mass_accuracy (MS2_precursor, 0, x.mass, Para_tolerance_precursor, Para_tolerance_precursor_unit) or Func_mass_accuracy (MS2_precursor, c13_precursor_shift, x.mass, Para_tolerance_precursor, Para_tolerance_precursor_unit)) and MS2_diag_check and x.count > 1:
				MS2_precursor_composition = x.composition
				MS2_precursor_description = x.description
				j = j_indexstore
				MS2_c13 = 0

				MS2_seq_cov_data = []
				for i in range (x.count):
					MS2_seq_cov_data.append (False)

                ### matches C13 peak
				if Func_mass_accuracy (MS2_precursor, c13_precursor_shift, x.mass, Para_tolerance_precursor, Para_tolerance_precursor_unit):
					MS2_precursor -= c13_precursor_shift
					MS2_c13 = 1 

                ### set a variable to calculate total abundance for this run
				MS2_intensity_sum = 0.0001
				MS2_intensity_sum_all = 0.0
				MS2_intensity_sum_real = 0.0
				MS2_intensity_sum_false = 0.0
				MS2_intensity_max = 0.0
				MS2_cov_int = 0.0
				MS2_cov_seq = 0.0

				MS2_real_mass_exp = []
				MS2_real_abundance = []
				MS2_real_mass_theo = []
				MS2_real_desc = []
				MS2_real_comp = []

				MS2_false_mass_exp = []
				MS2_false_abundance = []
				MS2_false_mass_theo = []
				MS2_false_desc = []
				MS2_false_comp = []

				while 'IONS' not in File_array[j] and '=' not in File_array[j]:
					MS2_peak_now = File_array[j].split (' ')
					MS2_peak_now_mz = float (MS2_peak_now[0])
					MS2_peak_now_inten = float (MS2_peak_now[1])

                    ### total intensity for MS/MS ions
					MS2_intensity_sum_all += MS2_peak_now_inten

                    ### highest intenstity of all MS/MS ions
					if MS2_peak_now_inten > MS2_intensity_max:
						MS2_intensity_max = MS2_peak_now_inten

                    ### first match peaks with real fragments
					for i in range(1, Test_charge_max + 1):
						for y in (MS2_fragment_list[int (MS2_peak_now_mz)-1] + MS2_fragment_list[int (MS2_peak_now_mz)] + MS2_fragment_list[int (MS2_peak_now_mz) + 1]):
							if y.charge == i and y.charge <= MS2_charge and Func_mass_accuracy (MS2_peak_now_mz, 0, y.mass, Para_tolerance_fragment, Para_tolerance_fragment_unit) and MS2_peak_now_inten > 0:
								MS2_fragment_composition = y.composition
								if Func_composition_inclusion (MS2_precursor_composition, MS2_fragment_composition):
									MS2_real_mass_exp.append (File_array[j].split (' ')[0])
									MS2_real_abundance.append (MS2_peak_now_inten)
									MS2_real_mass_theo.append (y.mass)
									MS2_real_desc.append (y.description + ', ' + str (y.charge) + ' + ')
									MS2_real_comp.append (y.composition)
									MS2_intensity_sum_real += MS2_peak_now_inten
									MS2_peak_now_inten = 0
									MS2_seq_cov_data[y.count - 1] = True

                    ### this advances iterator inside each test chunk
					j += 1

				MS2_intensity_sum += MS2_intensity_sum_real
				MS2_intensity_sum += MS2_intensity_sum_false

                ### output real fragments
				if MS2_c13 == 1:
					MS2_precursor += c13_precursor_shift

				MS2_seq_cov = 0.0
				for i in range (x.count - 1):
					if MS2_seq_cov_data[i] == True:
						MS2_seq_cov += 1

				MS2_cov_int = int (MS2_intensity_sum_real / MS2_intensity_sum_all * 100)
				MS2_cov_seq = int (MS2_seq_cov / (x.count - 1) * 100)

                ### output result for current glycan candidate
				if Func_FDR_check (MS2_precursor, MS2_intensity_sum_real / MS2_intensity_sum_all, FDR_final_int) and Func_FDR_check (MS2_precursor, MS2_seq_cov / (x.count - 1), FDR_final_seq): 
					if MS2_title not in MS2_identified_title:
						MS2_identified_title.append (MS2_title)
						MS2_identified_current_score = MS2_cov_int + MS2_cov_seq
						MS2_identified_list.append (str (MS2_title) + str (x.composition))
					if MS2_title in MS2_identified_title and MS2_cov_int + MS2_cov_seq - MS2_identified_current_score > 0:
						MS2_identified_current_score = MS2_cov_int + MS2_cov_seq
						MS2_identified_list.pop(-1)
						MS2_identified_list.append (str (MS2_title) + str (x.composition))
					
                ### print '  spectrum:\t', MS2_title, '\n  combination:\t', MS2_precursor_description, '\n'
	j += 1

MS2_precursor_list = Func_index_database (Func_combinations (MS2_ingredient_array, MS2_precursor_mass_max, 1, Para_NGlycan), Para_charged_mass)
MS2_fragment_list = Func_index_database (Func_combinations (MS2_fragment_array, MS2_precursor_mass_max, Test_charge_annotation), Para_charged_mass)

print ('\nprocessing annotation spectra')
MS2_spectra_no_now = 0

j = 0
while j < len (File_array):

    ### nonassignable list
	if MS2_precursor_description == 'a' and MS2_printed:
		MS2_nonassignable.append (str (MS2_title) + '\t' + str (MS2_pepmass) + '\t' + str (MS2_rt) + '\tDoes not fit any combination\t\t' + str (MS2_charge) + '\t\n')
		MGF_list_nonassignable.append (MS2_title)
		MS2_printed = False

	if 'BEGIN' in File_array[j]:
        ### progress of spectral processing
		MS2_spectra_no_now += 1
		if (MS2_spectra_no_now % 1000 == 0 or MS2_spectra_no_now == MS2_spectra_no) and not Test_output_spe:
			print ('Processing spectra: # ', MS2_spectra_no_now, ' / total', MS2_spectra_no)

		if (MS2_spectra_no_now % 100 == 0 or MS2_spectra_no_now == MS2_spectra_no) and Test_output_spe:
			print ('Processing spectra: # ', MS2_spectra_no_now, ' / total', MS2_spectra_no)

        ### reset the diagnostic ions
		for i in range (len (MS2_diag_exist)):
			MS2_diag_exist[i] = 0
			MS2_diag_exist_orig[i] = 0

        ### spectral general information
		MS2_title = str (File_array[j + 1].split ('=')[1])
		MS2_charge = int (File_array[j + 2].split ('=')[1][0])
		MS2_rt = float ('%.3f' % ((float (File_array[j + 3].split('=')[1]))/60))
		MS2_pepmass = float ('%.4f' % (float (File_array[j + 4].split('=')[1])))
		MS2_precursor = MS2_pepmass * float (MS2_charge) - (MS2_charge-1) * Para_charged_mass
		MS2_precursor_description = 'a'
		MS2_printed = True

		j += 5

        ### store original array position incase we need to look at the same peptide multiple times
		j_indexstore = j

		while 'IONS' not in File_array[j] and '=' not in File_array[j]:
			MS2_peak_now = File_array[j].split (' ')

            ### generate the diagnostic information
			for i in range (len (MS2_diag_exist)):
				if Func_mass_accuracy (MS2_peak_now[0], Para_charged_mass, MS2_diag_mass[i], Para_tolerance_fragment, Para_tolerance_fragment_unit):
					MS2_diag_exist[i] = 1
				if Func_mass_accuracy (MS2_peak_now[0], Para_charged_mass, MS2_diag_mass_orig[i], Para_tolerance_fragment, Para_tolerance_fragment_unit):
					MS2_diag_exist_orig[i] = 1
			j += 1

        ### generate combinations for every precursor value
		for x in (MS2_precursor_list[int (MS2_precursor)-1] + MS2_precursor_list[int (MS2_precursor)] + MS2_precursor_list[int (MS2_precursor) + 1]):
			MS2_diag_check = True

            ### check the diagnostic information
			for i in range (len (MS2_diag_exist)):
				if MS2_diag_exist[i] == 0 and MS2_diag_desc[i] in x.description:
					MS2_diag_check = False
				if MS2_diag_exist[i] == 1 and MS2_diag_exist_orig[i] == 1 and MS2_diag_desc[i] not in x.description:
					MS2_diag_check = False

            ### find precursor combination
			if (Func_mass_accuracy (MS2_precursor, 0, x.mass, Para_tolerance_precursor, Para_tolerance_precursor_unit) or Func_mass_accuracy (MS2_precursor, c13_precursor_shift, x.mass, Para_tolerance_precursor, Para_tolerance_precursor_unit)) and MS2_diag_check and x.count > 1:
				MS2_precursor_composition = x.composition
				MS2_precursor_description = x.description
				j = j_indexstore
				MS2_c13 = 0
				File_result_fragment.append ('spectrum:\t' + MS2_title + '\n')

				MS2_seq_cov_data = []
				for i in range (x.count):
					MS2_seq_cov_data.append (False)

                ### matches C12 peak
				if Func_mass_accuracy (MS2_precursor, 0, x.mass, Para_tolerance_precursor, Para_tolerance_precursor_unit):
					File_result_fragment.append ('precursor:\t' + str (MS2_precursor) + '\n')

                ### matches C13 peak
				if Func_mass_accuracy (MS2_precursor, c13_precursor_shift, x.mass, Para_tolerance_precursor, Para_tolerance_precursor_unit):
					MS2_precursor -= c13_precursor_shift
					MS2_c13 = 1 
					File_result_fragment.append ('precursor:\t' + str (MS2_precursor + c13_precursor_shift) + '\n')

                ### set a variable to calculate total abundance for this run
				MS2_intensity_sum = 0.0001
				MS2_intensity_sum_all = 0.0
				MS2_intensity_sum_real = 0.0
				MS2_intensity_sum_false = 0.0
				MS2_intensity_max = 0.0
				MS2_intensity_max_300 = 0.0
				MS2_cov_int = 0.0
				MS2_cov_seq = 0.0

				MS2_real_mass_exp = []
				MS2_real_abundance = []
				MS2_real_mass_theo = []
				MS2_real_desc = []
				MS2_real_comp = []
				MS2_real_annotation = []
				
				MS2_false_mass_exp = []
				MS2_false_abundance = []
				MS2_false_mass_theo = []
				MS2_false_desc = []
				MS2_false_comp = []

                ### heading for each glycan candidate
				File_result_fragment.append (Para_composition_glycan + '\t' + MS2_precursor_composition + '\n')
				File_result_fragment.append ('description:\t' + MS2_precursor_description + '\n')
				File_result_fragment.append ('retention:\t' + str (MS2_rt) + '\n')
				File_result_fragment.append ('charge:\t' + str (MS2_charge) + '\n')
				File_result_fragment.append ('Mass_matched\tMass_theroetical\tPPM\tAbundance\tComposition: ' + Para_composition_glycan + '\tDescription' + '\n')

				while 'IONS' not in File_array[j] and '=' not in File_array[j]:
					MS2_peak_now = File_array[j].split (' ')
					MS2_peak_now_mz = float (MS2_peak_now[0])
					MS2_peak_now_inten = float (MS2_peak_now[1])
					Spe_plot_allpeak_mass.append (MS2_peak_now_mz)
					Spe_plot_allpeak_abundance.append (MS2_peak_now_inten)

                    ### total intensity for MS/MS ions
					MS2_intensity_sum_all += MS2_peak_now_inten

                    ### highest intenstity of all MS/MS ions
					if MS2_peak_now_inten > MS2_intensity_max:
						MS2_intensity_max = MS2_peak_now_inten
					if MS2_peak_now_mz > 300 and MS2_peak_now_inten > MS2_intensity_max_300:
						MS2_intensity_max_300 = MS2_peak_now_inten

                    ### first match peaks with real fragments
					for i in range(1, Test_charge_annotation + 1):
						for y in (MS2_fragment_list[int (MS2_peak_now_mz)-1] + MS2_fragment_list[int (MS2_peak_now_mz)] + MS2_fragment_list[int (MS2_peak_now_mz) + 1]):
							if y.charge == i and y.charge <= MS2_charge and Func_mass_accuracy (MS2_peak_now_mz, 0, y.mass, Para_tolerance_fragment, Para_tolerance_fragment_unit) and MS2_peak_now_inten > 0:
								MS2_fragment_composition = y.composition
								if Func_composition_inclusion (MS2_precursor_composition, MS2_fragment_composition):
									MS2_real_mass_exp.append (float (File_array[j].split (' ')[0]))
									MS2_real_abundance.append (MS2_peak_now_inten)
									MS2_real_mass_theo.append (float (y.mass))
									MS2_real_desc.append (y.description + ', ' + str (y.charge) + ' + ')
									MS2_real_annotation.append (y.description + ' (' + str (y.count) + '): ' + str (y.charge) + '+, ' + str ('%.4f'% (MS2_peak_now_mz)))
									MS2_real_comp.append (y.composition)
									MS2_intensity_sum_real += MS2_peak_now_inten
									MS2_peak_now_inten = 0
									MS2_seq_cov_data[y.count - 1] = True

                    ### unmatched peaks will be matched with false fragments
					if MS2_peak_now_inten > 0 and 'yes' in Para_temfalse:
						for u in (MS2_precursor_list[int (MS2_precursor)-1] + MS2_precursor_list[int (MS2_precursor)] + MS2_precursor_list[int (MS2_precursor) + 1]):
							MS2_diag_check = True

                            ### check the diagnostic information
							for i in range (len (MS2_diag_exist)):
								if MS2_diag_exist[i] == 0 and MS2_diag_desc[i] in u.description:
									MS2_diag_check = False
								if MS2_diag_exist[i] == 1 and MS2_diag_exist_orig[i] == 1 and MS2_diag_desc[i] not in x.description:
									MS2_diag_check = False

                            ### find precursor combination
							if (Func_mass_accuracy (MS2_precursor, 0, u.mass, Para_tolerance_precursor, Para_tolerance_precursor_unit) or Func_mass_accuracy (MS2_precursor, c13_precursor_shift, u.mass, Para_tolerance_precursor, Para_tolerance_precursor_unit)) and MS2_diag_check:
								MS2_precursor_false_composition = u.composition

                                ### find fragments match
								for i in range(1, Test_charge_annotation + 1):
									for v in (MS2_fragment_list[int (MS2_peak_now_mz)-1] + MS2_fragment_list[int (MS2_peak_now_mz)] + MS2_fragment_list[int (MS2_peak_now_mz) + 1]):
										if v.charge == i and v.charge <= MS2_charge and Func_mass_accuracy (MS2_peak_now_mz, 0, v.mass, Para_tolerance_fragment, Para_tolerance_fragment_unit) and MS2_peak_now_inten > 0:
											MS2_fragment_composition = v.composition

											if Func_composition_inclusion (MS2_precursor_false_composition, MS2_fragment_composition) and MS2_peak_now_inten > 0:
												MS2_false_mass_exp.append (float (File_array[j].split (' ')[0]))
												MS2_false_abundance.append (MS2_peak_now_inten)
												MS2_false_mass_theo.append (float (v.mass))
												MS2_false_desc.append (v.description + ', ' + str (v.charge) + ' + ')
												MS2_false_comp.append (v.composition)
												MS2_intensity_sum_false += MS2_peak_now_inten
												MS2_peak_now_inten = 0

                            ### if fragment match found, stop
							if MS2_peak_now_inten == 0:
								break

                    ### this advances iterator inside each test chunk
					j += 1

				MS2_intensity_sum += MS2_intensity_sum_real
				MS2_intensity_sum += MS2_intensity_sum_false

                ### output real fragments
				for i,r in enumerate (MS2_real_mass_exp):
					File_result_fragment.append (('%.4f'% (MS2_real_mass_exp[i])) + '\t' + ('%.4f'% (MS2_real_mass_theo[i])) + '\t' + ('%.1f'% ((MS2_real_mass_exp[i] - MS2_real_mass_theo[i]) / MS2_real_mass_theo[i] * 1000000)) + '\t' + ('%.3f'% (float (MS2_real_abundance[i]))) + '\t' + MS2_real_comp[i] + '\t' + MS2_real_desc[i] + '\n')

				File_result_fragment.append ('Mass_matched_total:\t' + ('%.3f'% (float (MS2_intensity_sum_real))) + '\n')
 
                ### output false fragments
				if 'yes' in Para_temfalse:
                    # if I haven't seen any false
					if not MS2_false_mass_exp:
						File_result_fragment.append ( 'no false' + '\n')
					else:
						File_result_fragment.append ( '(falselist)' + '\n')

						for i,r in enumerate(MS2_false_mass_exp):
							File_result_fragment.append (('%.4f'% float (MS2_false_mass_exp[i])) + '\t' + ('%.4f'% (float (MS2_false_abundance[i]))) + '\t' + ('%.3f'% (float (float (MS2_false_abundance[i])/(float (MS2_intensity_sum))))) + '\t' + ('%.3f'% (MS2_false_mass_theo[i])) + '\t' + MS2_false_comp[i] + '\t' + MS2_false_desc[i] + '\n')

						File_result_fragment.append ('ab.false.total:\t' + ('%.3f'% (float (MS2_intensity_sum_false))) + '\n')

				if MS2_c13 == 1:
					MS2_precursor += c13_precursor_shift

				MS2_seq_cov = 0.0
				for i in range (x.count - 1):
					if MS2_seq_cov_data[i] == True:
						MS2_seq_cov += 1

				MS2_cov_int = int (MS2_intensity_sum_real / MS2_intensity_sum_all * 100)
				MS2_cov_seq = int (MS2_seq_cov / (x.count - 1) * 100)

                ### output result for current glycan candidate
				if (str (MS2_title) + str (x.composition)) in MS2_identified_list: 
					File_result.append (str (MS2_title) + '\t' + str (MS2_rt) + '\t' + str (MS2_precursor) + '\t' + str ('%.4f'% (x.mass)) + '\t' + str ('%.1f'% (float ((MS2_precursor - MS2_c13 * c13_precursor_shift - x.mass) / x.mass * 1000000))) + '\t' + str (MS2_precursor_composition) + '\t' + str (MS2_precursor_description) + '\t' + str (MS2_charge) + '\t' + str (MS2_cov_int) + '%\t' + str (MS2_cov_seq) + '%\t' + str (MS2_c13) + '\t' + str ('%.3f'% (MS2_intensity_sum_all)) + '\n')
                    #File_result.append (str (MS2_title) + '\t' + str (MS2_rt) + '\t' + str (MS2_precursor) + '\t' + str (x.mass) + '\t' + str ('%.1f'% (float ((MS2_precursor - MS2_c13 * c13_precursor_shift - x.mass) / x.mass * 1000000))) + '\t' + str (MS2_precursor_composition) + '\t' + str (MS2_precursor_description) + '\t' + str (MS2_charge) + '\t' + str ('%.3f'% (MS2_intensity_sum_real / MS2_intensity_sum_all)) + '\t' + str ('%.3f'% (MS2_seq_cov / (x.count - 1))) + '\t' + str (MS2_c13) + '\t' + str (MS2_intensity_sum_all) + '\n')

				File_result_fragment.append ('Coverage of intensity\t' + str (MS2_cov_int) + '%\n')
				File_result_fragment.append ('Coverage of sequence\t' + str (MS2_cov_seq) + '%\n')
				File_result_fragment.append ( '\n\n\n' )

                ### Main polt program
                ### size of figure and font
				Spe_plt.rcParams['font.size'] = 20
				Spe_plt.rcParams['figure.figsize'] = 36,18
				Spe_plt.rcParams['figure.dpi'] = 100

                ### get the correct abundances for each tested peak out of 100
				for i,r in enumerate (Spe_plot_allpeak_mass):
					if float (Spe_plot_allpeak_abundance[i]) < MS2_intensity_max_300:
						Spe_plot_allpeak_abundance[i] = float (Spe_plot_allpeak_abundance[i]) / float (MS2_intensity_max)
					if float (Spe_plot_allpeak_abundance[i]) >= MS2_intensity_max_300:
						Spe_plot_allpeak_abundance[i] = float (MS2_intensity_max_300) / float (MS2_intensity_max)

                ### remove all peaks that are bigger than the precursor mass
				for i,r in enumerate (Spe_plot_allpeak_mass):
					if Spe_plot_allpeak_mass[i] > (1 * float (MS2_precursor)):
						Spe_plot_allpeak_abundance [i] = 0

                ### peak information for annotation
				for i,r in enumerate (MS2_real_mass_exp):
					Spe_x_list.append (float (MS2_real_mass_exp[i])) 
					if float (MS2_real_abundance[i]) < MS2_intensity_max_300:
						Spe_y_list.append (float (MS2_real_abundance[i]) / float (MS2_intensity_max)) 
					if float (MS2_real_abundance[i]) >= MS2_intensity_max_300:
						Spe_y_list.append (float (MS2_intensity_max_300) / float (MS2_intensity_max)) 

                    ### get annotations on real peaks
					MS2_real_annotation[i] += (', ' + str ('%.3f'% (float (MS2_real_abundance[i]) / float (MS2_intensity_max))))
					Spe_annotate_list.append (MS2_real_annotation[i])
					Spe_annotated = []

                #### set the score threshold for plotting
				if (str (MS2_title) + str (x.composition)) in MS2_identified_list and Test_output_spe: 
					MGF_list_annotated.append (MS2_title)
					Spe_fig = Spe_plt.figure (num=1)
					Spe_ax1 = Spe_fig.add_subplot (111)

                    ### all peaks in black
					Spe_ax1.bar (Spe_plot_allpeak_mass, Spe_plot_allpeak_abundance, color='#000000', edgecolor='none', width=(0.003 * (max (Spe_x_list))))

                    ### matched peaks in red
					Spe_ax1.bar (Spe_x_list, Spe_y_list, color='#FF0000', edgecolor='none', width=(0.004 * (max (Spe_x_list))))
					Spe_ax1.set_xlabel ('M/Z')
					Spe_ax1.set_ylabel ('Abundance')
					Spe_ax1.set_title ('Title: ' + str (MS2_title) + '      RT: ' + str (MS2_rt) + '      Precursor: ' + str (MS2_precursor) + '\nComposition: ' + str (MS2_precursor_composition) + ' (' + str (Para_composition_glycan) + ')\nDescription: ' + str (MS2_precursor_description) + '\nCov. Int : ' + str (MS2_cov_int) + '%      Cov. Seq : ' + str (MS2_cov_seq) + '%' + '      Peak annotation by [description (# of glycans) : charge, m/z, rel. intensity]')
					Spe_ax1.set_ylim ([0, 2.2 * MS2_intensity_max_300 / MS2_intensity_max])
					Spe_ax1.set_xlim ([0, max (Spe_x_list) * 1.5])
					Spe_y_list_counter = 2.1 * MS2_intensity_max_300 / MS2_intensity_max
					k = 0

                    ### now we put annotations on all real peaks
					for i,r in enumerate (Spe_x_list):

						if len (set (Spe_annotate_list)) > 30:
							Spe_annotate_line = 27
							Spe_annotate_font_size = 16
						elif len (set (Spe_annotate_list)) > 20:
							Spe_annotate_line = 21
							Spe_annotate_font_size = 18
						else:
							Spe_annotate_line = 15
							Spe_annotate_font_size = 20

                        ### space out the annotations in a round of X times
						if k == Spe_annotate_line : 
							Spe_y_list_counter = 2.1 * MS2_intensity_max_300 / MS2_intensity_max
							k = 0

						if Spe_annotate_list[i] not in Spe_annotated:
							Spe_ax1.annotate (Spe_annotate_list[i], xy=(Spe_x_list[i], Spe_y_list[i]), xytext=(Spe_x_list[i], Spe_y_list_counter), arrowprops=dict(linestyle='dotted', facecolor='none', edgecolor='none', color='#bbbbbb', shrink=0.1, width=0.1, headwidth=0.1), rotation='horizontal', fontsize=Spe_annotate_font_size)
							Spe_annotated.append (Spe_annotate_list[i])
							Spe_y_list_counter += -(1.1 / float (Spe_annotate_line)) * MS2_intensity_max_300 / MS2_intensity_max
							k += 1

                    ### print ('spectrum plotted')
					Spe_plt.savefig(str (MS2_title.split (',')[0]) + ' RT=' + str (MS2_rt) + ' GLY=' + str (MS2_precursor_description) + '  Mass=' + str (MS2_pepmass) + ' Z=' + str (MS2_charge) + '.png')
				
				Spe_x_list = []
				Spe_y_list = []
				Spe_annotate_list = []
				Spe_annotated = []
				Spe_plot_allpeak_mass = []
				Spe_plot_allpeak_abundance = []
				Spe_plt.clf()

                # print '  spectrum:\t', MS2_title, '\n  combination:\t', MS2_precursor_description, '\n'
	j += 1

### output library

Output_library_rt_max = 0
for i in range (1, len (File_result)):
	if int (float (File_result[i].split ('\t')[1])) > Output_library_rt_max:
		Output_library_rt_max = int (float (File_result[i].split ('\t')[1]))

Output_library_rt = []
Output_library_mass = []
Output_library_composition = []
Output_library_description = []
Output_library_TI = []

for i in range (1, len (File_result)):
	Output_library_rt.append (int (float (File_result[i].split ('\t')[1])))
	Output_library_mass.append (float ('%.4f'% float (File_result[i].split ('\t')[3])))
	Output_library_composition.append (File_result[i].split ('\t')[5])
	Output_library_description.append (File_result[i].split ('\t')[6])
	Output_library_TI.append (float ('%.3f'% float(File_result[i].split ('\t')[11])))

Output_library_description_set = list (set (Output_library_description))
Output_library_description_set.sort()

Output_library_composition_set = []
Output_library_mass_set = []
Output_library_TI_set = []

for i in range (len (Output_library_description_set)):
	Output_library_mass_set.append (0)
	Output_library_composition_set.append (0)
	Output_library_TI_set.append (0)

for i in range (len (Output_library_description_set)):
	for j in range (len (Output_library_description)):
		if Output_library_description_set[i] == Output_library_description[j]:
			Output_library_mass_set[i] = Output_library_mass[j]
			Output_library_composition_set[i] = Output_library_composition[j]
			Output_library_TI_set[i] += Output_library_TI[j]
	Output_library_TI_set[i] = ('%.3f'% float (Output_library_TI_set[i]))

Output_library = []
for i in range (len (Output_library_description_set) + 1):
	Output_library.append ([])

Output_library[0].append (Para_composition_glycan)
Output_library[0].append ('Description')
Output_library[0].append ('Glycan mass')
Output_library[0].append ('Total MS/MS intensity')
for i in range (Output_library_rt_max + 1):
	Output_library[0].append (i)

Output_library_row = 0
for i in range (len (Output_library_description_set)):
	Output_library_row += 1
	Output_library[Output_library_row].append (Output_library_composition_set[i])
	Output_library[Output_library_row].append (Output_library_description_set[i])
	Output_library[Output_library_row].append (Output_library_mass_set[i])
	Output_library[Output_library_row].append (Output_library_TI_set[i])

	for j in range (Output_library_rt_max + 1):
		Output_library_data_now = False
		Output_library_data = 0.0
		for k in range (len (Output_library_description)):
			if Output_library_description[k] == Output_library_description_set[i] and Output_library_rt[k] == j:
				Output_library_data_now = True
				Output_library_data += Output_library_TI[k]
		if Output_library_data_now:
			Output_library[Output_library_row].append ('%.3f'% float (Output_library_data))
		else:
			Output_library[Output_library_row].append ('')

Output_library_file_TI = list (reversed (sorted (list (set (Output_library_TI_set)))))

Output_library_file =[]
Output_library_file_now = ''

for i in range (len (Output_library[0])):
	Output_library_file_now += str (Output_library[0][i])
	Output_library_file_now += '\t'
Output_library_file_now += '\n'
Output_library_file.append (Output_library_file_now)

for i in range (len (Output_library_file_TI)):
	for j in range (1, len (Output_library)):
		if Output_library_file_TI[i] == Output_library[j][3]:
			Output_library_file_now = ''
			for k in range (len (Output_library[j])):
				Output_library_file_now += str (Output_library[j][k])
				Output_library_file_now += '\t'
			Output_library_file_now += '\n'
			Output_library_file.append (Output_library_file_now)

###
### result output

if Test_output_spe:
	os.chdir (os.pardir)

if Para_filter_diagnostic_ions == 'yes' and Test_output_dia_MGF:
	print ('\noutput MS/MS spectra with diagnostic ions')
	with open (Para_filename.split ('.')[0] + '-diagnostic-ions.mgf', 'w') as Output_MGF_diagnostic_ions:
		MGF__output = Func_filter_MGF (Para_filename, Para_filter_sn, Para_filter_top, MGF_list_diagnostic_ions)
		for line in MGF__output:
			Output_MGF_diagnostic_ions.write (line)

'''

### output MGF with annotated spectra
with open (Para_filename.split ('.')[0] + '-annotated.MGF', 'w') as Output_MGF_annotated:
	MGF__output = Func_filter_MGF (Para_filename, Para_filter_sn, Para_filter_top, MGF_list_annotated)
	for line in MGF__output:
		Output_MGF_annotated.write (line)

### output MGF with unassigned spectra
with open (Para_filename.split ('.')[0] + '-nonassignable.MGF', 'w') as Output_MGF_nonassignable:
	MGF__output = Func_filter_MGF (Para_filename, Para_filter_sn, Para_filter_top, MGF_list_nonassignable)
	for line in MGF__output:
		Output_MGF_nonassignable.write (line)
'''

Excel_output = Workbook()

Excel_output.create_sheet (index=0, title='Spectra')
Excel_output.create_sheet (index=1, title='Library')
Excel_output.create_sheet (index=2, title='Fragments')
Excel_output.create_sheet (index=3, title='Parameters')

Excel_output_data = Excel_output.get_sheet_by_name('Spectra')
for line in File_result:
	Excel_output_temp = line.split ('\t')
	Excel_output_data.append (Excel_output_temp)

Excel_output_data = Excel_output.get_sheet_by_name('Library')
for line in Output_library_file:
	Excel_output_temp = line.split ('\t')
	Excel_output_data.append (Excel_output_temp)

Excel_output_data = Excel_output.get_sheet_by_name('Fragments')
for line in File_result_fragment:
	Excel_output_temp = line.split ('\t')
	Excel_output_data.append (Excel_output_temp)

Excel_output_data = Excel_output.get_sheet_by_name('Parameters')
for line in File_array_parameter:
	Excel_output_temp = line.split ('\t')
	Excel_output_data.append (Excel_output_temp)

Excel_output.save(filename = Para_filename.split ('.')[0] + '-Glyconote-Result-MS2.xlsx')

###
### Potential glycan composition

print ('\n6th step - analysis of potential glycan composition\n(Diectory)', OS_dirname, '\n(Filename)', OS_filename, '\n')

'''
Data_file_result = []
Data_file_result_composition = []
for i in range (1, len (File_result)):
    Data_file_result.append (File_result[i].split('\t')[0] + '\t' + File_result[i].split('\t')[2])
    Data_file_result_composition.append (File_result[i].split('\t')[0] + '\t' + File_result[i].split('\t')[6])

Data_file_result = list (set (Data_file_result))

Data_composition_analysis_mass = []
Data_composition_analysis_title = []
for i in range (len (Data_file_result)):
    Data_composition_analysis_mass.append (float (Data_file_result[i].split('\t')[1]))
    Data_composition_analysis_title.append (Data_file_result[i].split('\t')[0])

mono_sac_name = []
mono_sac_mass = []
mono_sac_list = []
for x in MS2_ingredient_array:
    if 'RED' not in x.name and 'ADD' not in x.name and 'H2O' not in x.name:
        mono_sac_name.append (x.name)
        mono_sac_mass.append (x.mass)
        mono_sac_list.append (0)

for i in range (len (Data_composition_analysis_mass)):
    for j in range (len (Data_composition_analysis_mass)):
        for k in range (len (mono_sac_mass)):
            if Func_mass_accuracy (Data_composition_analysis_mass[i], mono_sac_mass[k], Data_composition_analysis_mass[j], Para_tolerance_precursor, Para_tolerance_precursor_unit) and Func_check_glycan (Data_composition_analysis_title[i], mono_sac_name[k], Data_file_result_composition):
                mono_sac_list[k] += 1

mono_sac_total = 0.0
for i in range (len (mono_sac_mass)):
    mono_sac_total += mono_sac_list[i]

with open (Para_filename.split ('.')[0] + '-open-search.txt', 'w') as Output_composition_analysis:
    for i in range (len (mono_sac_mass)):
        print mono_sac_name[i], '\t', ('%.3f'% (float (mono_sac_list[i]) / mono_sac_total))
        Output_composition_analysis.write (str (mono_sac_name[i]) + '\t' + str (mono_sac_mass[i]) + '\t' + str (mono_sac_list[i]) + '\t' + str ('{0:.0f}%'.format(float (mono_sac_list[i]) / mono_sac_total * 100)) + '\n')
'''

print ('\nSpectral analysis finished for file\n(Diectory)', OS_dirname, '\n(Filename)', OS_filename, '\n(Parameter)', Parameter_filename, '\n')

### finish greetings and total running time
messagebox.showinfo (title='Greetings', message='Spectral analysis finished in ' + ('--- %.0fs seconds' % (time.time() - Time_start)) + '\nfor file ' + Para_filename)